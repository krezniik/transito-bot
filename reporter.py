"""
Módulo de reportes.
Genera el resumen de texto en el mismo formato del script original,
envía al grupo de Telegram y exporta a Excel.
"""

import tempfile
from collections import defaultdict
from datetime import date, datetime
from typing import List, Dict
from telegram import Bot
from telegram.constants import ParseMode
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Resumen de texto (mismo formato que el script original) ──────────────────
LINEA_MAQUINA = {
    "Mespack 1": "L36",
    "Mespack 2": "L37",
    "Mespack 3": "L39",
    "Chub":      "L46",
}


def generar_resumen_texto(lotes: List[Dict]) -> str:
    """
    Genera el bloque "Tránsito 📋" agrupado por producto + presentación + mercado,
    igual al formato original del script.
    """
    resumen_limpio: dict = defaultdict(float)
    resumen_maquinas: dict = defaultdict(lambda: defaultdict(lambda: {"cajas": 0.0, "cpc": 0.0}))

    for r in lotes:
        clave = f"{r['producto_legible']} {r['presentacion'].lower()} {r['mercado_legible']}"
        resumen_limpio[clave]                          += r['cajas_en_transito']
        resumen_maquinas[r['maquina']][clave]["cajas"] += r['cajas_en_transito']
        resumen_maquinas[r['maquina']][clave]["cpc"]    = r['cajas_por_canasta']

    # ── Detalle por máquina ──
    lineas = ["⚙️ *Detalle por llenadora:*\n"]
    for maquina, productos in resumen_maquinas.items():
        total_maq = sum(d["cajas"] for d in productos.values())
        linea_num = LINEA_MAQUINA.get(maquina, "")
        encabezado = f"*{maquina} ({linea_num})*" if linea_num else f"*{maquina}*"
        lineas.append(encabezado)
        for clave, datos in productos.items():
            lineas.append(f"  {clave} — {datos['cpc']} c/c: {int(datos['cajas']):,} cajas")
        lineas.append(f"  _Total: {int(total_maq):,} cajas_\n")

    # ── Resumen consolidado (para compartir) ──
    lineas.append("───────────────────")
    lineas.append("Tránsito 📋\n")
    for clave, total in resumen_limpio.items():
        lineas.append(f"{clave}")
        lineas.append(f"{int(total):,} cajas 📦\n")

    total_general = sum(resumen_limpio.values())
    lineas.append(f"───────────────────")
    lineas.append(f"*Total general: {int(total_general):,} cajas*")

    return "\n".join(lineas)


def _generar_mensaje_grupo(lotes: List[Dict]) -> str:
    """Genera solo el bloque limpio para el grupo (sin detalle por máquina)."""
    resumen_limpio: dict = defaultdict(float)
    for r in lotes:
        clave = f"{r['producto_legible']} {r['presentacion'].lower()} {r['mercado_legible']}"
        resumen_limpio[clave] += r['cajas_en_transito']

    lineas = ["Tránsito 📋\n"]
    for clave, total in resumen_limpio.items():
        lineas.append(f"{clave}")
        lineas.append(f"{int(total):,} cajas 📦\n")
    return "\n".join(lineas)


async def enviar_reporte_grupo(bot: Bot, chat_id: str, lotes: List[Dict]):
    mensaje = _generar_mensaje_grupo(lotes)
    await bot.send_message(chat_id=chat_id, text=mensaje)


# ── Exportación a Excel ───────────────────────────────────────────────────────
def exportar_excel(lotes: List[Dict], desde: date, hasta: date) -> str:
    """
    Genera un archivo Excel con dos hojas:
      1. Detalle — cada lote como fila
      2. Resumen — agrupado por producto/presentación/mercado
    Devuelve la ruta del archivo temporal generado.
    """
    wb = openpyxl.Workbook()

    # ── Colores y estilos ──
    COLOR_HEADER    = "1F4E79"
    COLOR_SUBTOTAL  = "BDD7EE"
    COLOR_TOTAL     = "2E75B6"
    COLOR_ALT_ROW   = "EBF3FB"

    header_font    = Font(bold=True, color="FFFFFF", size=11)
    subtotal_font  = Font(bold=True, color="1F4E79", size=10)
    total_font     = Font(bold=True, color="FFFFFF", size=11)
    normal_font    = Font(size=10)
    center_align   = Alignment(horizontal="center", vertical="center")
    left_align     = Alignment(horizontal="left",   vertical="center")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def set_header_row(ws, row, values, bg_color=COLOR_HEADER):
        fill = PatternFill("solid", fgColor=bg_color)
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = header_font
            cell.fill      = fill
            cell.alignment = center_align
            cell.border    = thin_border

    def style_data_row(ws, row, n_cols, alt=False):
        fill = PatternFill("solid", fgColor=COLOR_ALT_ROW) if alt else None
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font   = normal_font
            cell.border = thin_border
            if fill:
                cell.fill = fill

    # ════════════════════════════════════════════════════
    # HOJA 1: DETALLE
    # ════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Detalle"

    titulo = f"Reporte de Tránsito — {desde} al {hasta}"
    ws1.merge_cells("A1:J1")
    title_cell = ws1["A1"]
    title_cell.value     = titulo
    title_cell.font      = Font(bold=True, size=13, color=COLOR_HEADER)
    title_cell.alignment = center_align
    ws1.row_dimensions[1].height = 25

    headers = ["ID Lote", "Fecha", "Hora", "Llenadora", "Canastas",
               "Cajas/Canasta", "Presentación", "Producto", "Mercado", "Cajas en Tránsito"]
    set_header_row(ws1, 2, headers)
    ws1.row_dimensions[2].height = 20

    for i, lote in enumerate(lotes):
        row_num = i + 3
        dt      = datetime.fromisoformat(lote["timestamp"])
        fecha   = dt.date().isoformat()
        hora    = dt.strftime("%H:%M")
        alt     = (i % 2 == 1)

        ws1.cell(row=row_num, column=1,  value=lote["id"])
        ws1.cell(row=row_num, column=2,  value=fecha)
        ws1.cell(row=row_num, column=3,  value=hora)
        ws1.cell(row=row_num, column=4,  value=lote["maquina"])
        ws1.cell(row=row_num, column=5,  value=lote["canastas"])
        ws1.cell(row=row_num, column=6,  value=lote["cajas_por_canasta"])
        ws1.cell(row=row_num, column=7,  value=lote["presentacion"])
        ws1.cell(row=row_num, column=8,  value=lote["producto_legible"].replace("🫘", "").strip())
        ws1.cell(row=row_num, column=9,  value=lote["mercado_legible"].replace("🇬🇹", "RTCA").replace("🇺🇸", "FDA"))
        ws1.cell(row=row_num, column=10, value=int(lote["cajas_en_transito"]))
        style_data_row(ws1, row_num, 10, alt)

    # Total
    total_row = len(lotes) + 3
    ws1.cell(row=total_row, column=9,  value="TOTAL").font = total_font
    ws1.cell(row=total_row, column=10, value=int(sum(l["cajas_en_transito"] for l in lotes)))
    fill_total = PatternFill("solid", fgColor=COLOR_TOTAL)
    for col in range(1, 11):
        cell = ws1.cell(row=total_row, column=col)
        cell.fill   = fill_total
        cell.font   = total_font
        cell.border = thin_border
        cell.alignment = center_align

    # Ancho de columnas
    col_widths = [10, 12, 8, 14, 10, 14, 14, 22, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.freeze_panes = "A3"

    # ════════════════════════════════════════════════════
    # HOJA 2: RESUMEN CONSOLIDADO
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Resumen")

    ws2.merge_cells("A1:E1")
    tc2 = ws2["A1"]
    tc2.value     = titulo
    tc2.font      = Font(bold=True, size=13, color=COLOR_HEADER)
    tc2.alignment = center_align
    ws2.row_dimensions[1].height = 25

    set_header_row(ws2, 2, ["Llenadora", "Producto", "Presentación", "Mercado", "Total Cajas"])
    ws2.row_dimensions[2].height = 20

    resumen: dict = defaultdict(float)
    resumen_maq: dict = defaultdict(lambda: defaultdict(float))
    for l in lotes:
        clave = (l["maquina"], l["producto_legible"].replace("🫘","").strip(),
                 l["presentacion"], l["mercado_legible"].replace("🇬🇹","RTCA").replace("🇺🇸","FDA"))
        resumen[clave]                         += l["cajas_en_transito"]
        resumen_maq[l["maquina"]][clave[1:]]   += l["cajas_en_transito"]

    row_num = 3
    alt = False
    for (maq, prod, pres, merc), total in sorted(resumen.items()):
        ws2.cell(row=row_num, column=1, value=maq)
        ws2.cell(row=row_num, column=2, value=prod)
        ws2.cell(row=row_num, column=3, value=pres)
        ws2.cell(row=row_num, column=4, value=merc)
        ws2.cell(row=row_num, column=5, value=int(total))
        style_data_row(ws2, row_num, 5, alt)
        alt = not alt
        row_num += 1

    # Subtotales por máquina
    row_num += 1
    set_header_row(ws2, row_num, ["Subtotales por Llenadora", "", "", "", ""], COLOR_SUBTOTAL)
    ws2.cell(row=row_num, column=1).font = Font(bold=True, color="1F4E79")
    row_num += 1
    for maq, prods in resumen_maq.items():
        total_maq = sum(prods.values())
        ws2.cell(row=row_num, column=1, value=maq).font = subtotal_font
        ws2.cell(row=row_num, column=5, value=int(total_maq)).font = subtotal_font
        row_num += 1

    # Total general
    total_gral = int(sum(l["cajas_en_transito"] for l in lotes))
    for col in range(1, 6):
        cell = ws2.cell(row=row_num, column=col)
        cell.fill   = PatternFill("solid", fgColor=COLOR_TOTAL)
        cell.font   = total_font
        cell.border = thin_border
        cell.alignment = center_align
    ws2.cell(row=row_num, column=4, value="TOTAL GENERAL").font = total_font
    ws2.cell(row=row_num, column=5, value=total_gral).font      = total_font

    col_widths2 = [16, 22, 14, 10, 16]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    # ── Guardar ──
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
